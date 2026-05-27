"""Build the tidy long-format work-permits-by-occupation-group master CSV from
Migrationsverket's "Beviljade arbetstillstånd" xlsx/xls files.

Output goes to ./csv/master_work_permits_by_occupation_group.csv. Every Swedish
label is paired with an English column. Counts are in `granted_count`. Months are
'YYYY-MM' strings; annual-only rows have month == NULL and year_month == 'YYYY'.
"""
from __future__ import annotations
import csv, re, sys
from pathlib import Path
import openpyxl, xlrd

sys.path.insert(0, str(Path(__file__).parent))
import translations as T

BASE = Path("/Users/alejandro.lozadacort/the-local/the-local/migration-statistics")
OUT = BASE / "csv"
OUT.mkdir(exist_ok=True)


def clean(s):
    if s is None: return None
    if isinstance(s, str): return re.sub(r"\s+", " ", s).strip()
    return s


def tr(dct: dict, key, fallback_label="(unmapped)"):
    if key is None: return None
    k = clean(key)
    # Normalise some encoding artefacts
    k_norm = k.replace("m0m0", "m.m.")
    if k_norm in dct: return dct[k_norm]
    if k in dct: return dct[k]
    # case-insensitive fallback
    for sk, sv in dct.items():
        if sk.casefold() == k.casefold(): return sv
    # try stripping reference markers like "(4 KAP 2 UTLL)" or "1)" suffixes
    k2 = re.sub(r"\s*\d+\)\s*$", "", k).strip()
    if k2 in dct: return dct[k2]
    return f"{fallback_label}: {k}"


def normalize_month_header(h):
    """Accept '2024-01', datetime, etc. Return 'YYYY-MM' string or None for Totalt."""
    if h is None: return None
    if isinstance(h, str):
        s = h.strip()
        if s.lower() in ("totalt", "total"): return None
        m = re.match(r"^(\d{4})-(\d{1,2})$", s)
        if m: return f"{m.group(1)}-{int(m.group(2)):02d}"
        return s
    return str(h)


def to_int(x):
    if x is None or x == "": return None
    if isinstance(x, (int, float)):
        if isinstance(x, float) and x != x: return None
        return int(round(x))
    if isinstance(x, str):
        x = x.strip()
        if not x: return None
        m = re.match(r"^\s*(\d+)", x)
        if m: return int(m.group(1))
    return None


# ------------------------------------------------------------------
# Work permits — by occupation group (Yrkesgrupp)
# ------------------------------------------------------------------
def build_occupation_group():
    rows = []
    for f in sorted(BASE.glob("granted_permits_2015_2025/Beviljade*arbetstillstånd*.xls*")):
        year = int(re.search(r"(20\d{2})", f.name).group(1))
        if year <= 2020:
            rows.extend(_parse_occupation_group_legacy(f, year))
        else:
            rows.extend(_parse_occupation_group_modern(f, year))
    return rows


def _parse_occupation_group_modern(f, year):
    """2021 and later: openpyxl, monthly columns, first-time/extension split (2022+)."""
    rows = []
    wb = openpyxl.load_workbook(f, data_only=True)
    for sheet in wb.sheetnames:
        if "Yrkesgrupp" not in sheet: continue
        permit_type_sv = "förlängning" if "förlängning" in sheet else "förstagångs"
        ws_rows = list(ws_iter(wb[sheet]))
        header_row_idx, header = find_header(ws_rows, "Yrkesområde")
        if header is None: continue
        last_field = None
        for r in ws_rows[header_row_idx + 1:]:
            if r[0]: last_field = clean(r[0])
            grp = clean(r[1])
            if not grp: continue
            for ci, col_name in enumerate(header[2:], start=2):
                month = normalize_month_header(col_name)
                val = to_int(r[ci])
                if val is None: continue
                is_total = (month is None) and (str(col_name).strip().lower() == "totalt")
                rows.append({
                    "year": year,
                    "month": int(month.split("-")[1]) if month else None,
                    "year_month": month if month else str(year),
                    "is_year_total": is_total,
                    "permit_type_sv": permit_type_sv,
                    "permit_type_en": T.PERMIT_TYPE[permit_type_sv],
                    "occupation_field_sv": last_field,
                    "occupation_field_en": tr(T.YRKESOMRADE, last_field) if last_field else None,
                    "occupation_group_sv": grp,
                    "occupation_group_en": tr(T.YRKESGRUPP, grp),
                    "granted_count": val,
                })
    wb.close()
    return rows


def _parse_occupation_group_legacy(f, year):
    """2015-2020: annual totals only, first-time permits only ('Exklusive förlängningar'),
    single 'Yrkesområde, Yrkesgrupp' sheet. The sheet repeats a [Yrkesområde, _, Yrkesgrupp, Antal]
    header for every field block, ends each block with a 'Summa:' subtotal, and finishes with
    a 'TOTALT' grand-total row. Layout differs slightly between .xls and .xlsx but the column
    positions are stable: field=col 1, group=col 3, count=col 4."""
    if f.suffix.lower() == ".xls":
        wb = xlrd.open_workbook(f)
        sh = wb.sheet_by_name("Yrkesområde, Yrkesgrupp")
        ws_rows = [tuple(sh.cell_value(r, c) for c in range(sh.ncols)) for r in range(sh.nrows)]
    else:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws_rows = list(ws_iter(wb["Yrkesområde, Yrkesgrupp"]))
        wb.close()

    rows = []
    last_field = None
    permit_type_sv = "förstagångs"
    prev_was_data = False
    for r in ws_rows:
        # ensure 5 cells so col 1/3/4 access is safe
        cells = list(r) + [None] * (5 - len(r))
        field = clean(cells[1])
        grp = clean(cells[3])
        val = to_int(cells[4])
        # skip header / metadata / blank rows
        if field in ("Yrkesområde", "Rubrik", "Senast körd"):
            prev_was_data = False
            continue
        if field == "TOTALT":
            prev_was_data = False
            continue
        if grp in ("Yrkesgrupp",):
            prev_was_data = False
            continue
        # update running field tag when present
        if field:
            last_field = field
        # 2015 trails a grand 'Summa:' row after a blank line, separate from
        # any field block; skip it (legitimate field summas immediately follow
        # data rows).
        if grp == "Summa:" and not prev_was_data:
            prev_was_data = False
            continue
        if not grp:
            prev_was_data = False
            continue
        if val is None:
            prev_was_data = False
            continue
        prev_was_data = True
        # Map per-field 'Summa:' subtotals to the canonical 'Totalt' pseudo-group
        # used in the 2021+ master CSV, so blind yearly sums (without filtering
        # Total rows) match in magnitude across the full 2015-2025 range.
        if grp == "Summa:":
            grp_sv, grp_en = "Totalt", "Total"
        else:
            grp_sv, grp_en = grp, tr(T.YRKESGRUPP, grp)
        rows.append({
            "year": year,
            "month": None,
            "year_month": str(year),
            "is_year_total": True,
            "permit_type_sv": permit_type_sv,
            "permit_type_en": T.PERMIT_TYPE[permit_type_sv],
            "occupation_field_sv": last_field,
            "occupation_field_en": tr(T.YRKESOMRADE, last_field) if last_field else None,
            "occupation_group_sv": grp_sv,
            "occupation_group_en": grp_en,
            "granted_count": val,
        })
    return rows


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------
def ws_iter(ws):
    for row in ws.iter_rows(values_only=True):
        yield row


def find_header(ws_rows, first_col_label):
    """Find the row whose first cell == first_col_label. Return (idx, row)."""
    for i, r in enumerate(ws_rows):
        if clean(r[0]) == first_col_label:
            return i, [clean(x) for x in r]
    return -1, None


def write_csv(path, rows, columns):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote {path.name}: {len(rows)} rows")


def main():
    print("Building CSV files...")

    rows = build_occupation_group()
    write_csv(OUT / "master_work_permits_by_occupation_group.csv", rows, [
        "year", "month", "year_month", "is_year_total",
        "permit_type_sv", "permit_type_en",
        "occupation_field_sv", "occupation_field_en",
        "occupation_group_sv", "occupation_group_en",
        "granted_count",
    ])


if __name__ == "__main__":
    main()
